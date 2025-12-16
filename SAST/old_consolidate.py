#!/usr/bin/env python3
"""
Consolidate multiple CSV scan outputs into a single XLSX file.

- Scans directory: /components/1.5.4/output/  (change BASE_OUTPUT_DIR)
- Multiprocessing: parses CSV files in parallel
- Threading (per-worker): uses a small threadpool to process row chunks
- Output: single XLSX file with each CSV's summary and detailed rows appended

Output layout (in single sheet "Consolidated"):
Row example:
------------------------------------------------------------
| <FILENAME>                                              |
| Critical: X  High: Y  Medium: Z  Low: W                |
| CVEs | Severity | Jfrog Severity | CVSS v3 | Cwe | Fix Version |
| ...rows...
<next file block>
------------------------------------------------------------
"""

import os
import glob
import math
import logging
from pathlib import Path
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from typing import Dict, List, Tuple, Any

import pandas as pd

# ------------
# CONFIG
# ------------
BASE_OUTPUT_DIR = "/components/1.5.4/output"   # adjust as needed
OUTPUT_XLSX = "/components/1.5.4/consolidated_scan_results.xlsx"
NUM_PROCESSES = max(1, os.cpu_count() - 1)     # multiprocessing workers
THREADS_PER_WORKER = 4                         # internal threadpool per worker
CHUNK_SIZE = 5000                              # number of rows per chunk if using chunked read

# ------------
# LOGGING
# ------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(processName)s/%(threadName)s - %(message)s"
)
logger = logging.getLogger("consolidator")


# ------------
# HELPERS
# ------------
def normalize_cols(cols: List[str]) -> Dict[str, str]:
    """Map lowercased column names -> actual column names for case-insensitive matching."""
    return {c.lower().strip(): c for c in cols}


def safe_get(df_row: pd.Series, col_map: Dict[str, str], key: str) -> str:
    """Return string value for a possibly-missing column (case-insensitive)"""
    real = col_map.get(key.lower())
    if not real:
        return ""
    val = df_row.get(real, "")
    if pd.isna(val):
        return ""
    return str(val).strip()


# ------------
# WORKER: parse single CSV file
# ------------
def process_csv_file(path: str) -> Dict[str, Any]:
    """
    Parse CSV file and return a dict with:
      - filename
      - counts: {critical, high, medium, low}
      - details: list of rows [CVEs, Severity, Jfrog Severity, CVSS v3, Cwe, Fix Version]
    This function is designed to be invoked in a separate process.
    """
    logger.info(f"Worker start: {path}")
    filename = os.path.basename(path)

    # Try to read with pandas. Use dtype=str to avoid weird conversions.
    # If file is large, use chunks and process in threadpool for speed.
    try:
        # First read only header to discover columns
        df_head = pd.read_csv(path, nrows=0, sep=None, engine="python")
        col_map = normalize_cols(df_head.columns.tolist())
    except Exception as e:
        logger.warning(f"Failed to read header from {filename}: {e}. Trying full read fallback.")
        try:
            df_full = pd.read_csv(path, dtype=str, sep=None, engine="python", encoding="utf-8")
            col_map = normalize_cols(df_full.columns.tolist())
        except Exception as e2:
            logger.error(f"Could not read CSV {filename}: {e2}")
            return {"filename": filename, "counts": {}, "details": []}

    # target column keys we need (case-insensitive)
    # user provided possible column names in prompt; we map them.
    needed_keys = {
        "cves": ["cves", "cve", "cve(s)"],
        "severity": ["severity"],
        "jfrog_severity": ["jfrog severity", "jfrog_severity", "jfrogseverity"],
        "cvss_v3": ["cvss v3", "cvss_v3", "cvssv3"],
        "cwe": ["cwe"],
        "fix_version": ["fix version", "fix_version", "fixversion", "fix"]
    }

    # reverse lookup to map our canonical key -> actual column name in file (if present)
    mapped_cols = {}
    for canonical, candidates in needed_keys.items():
        found = None
        for cand in candidates:
            real = col_map.get(cand.lower())
            if real:
                found = real
                break
        mapped_cols[canonical] = found  # may be None

    # We'll attempt chunked reading if file is large to control memory.
    # Build results using threads to process chunks concurrently inside the process.
    details_acc: List[Tuple[str, str, str, str, str, str]] = []
    severity_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}

    def process_chunk(df_chunk: pd.DataFrame) -> Tuple[List[Tuple[str, str, str, str, str, str]], Dict[str, int]]:
        chunk_details = []
        local_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        cols_lower = normalize_cols(df_chunk.columns.tolist())

        # Update mapped_cols for this chunk if not detected earlier
        resolved = {}
        for k, _ in needed_keys.items():
            if mapped_cols.get(k):
                resolved[k] = mapped_cols[k]
            else:
                # try to find in chunk columns
                for cand in needed_keys[k]:
                    if cand.lower() in cols_lower:
                        resolved[k] = cols_lower[cand.lower()]
                        break
                else:
                    resolved[k] = None

        for _, row in df_chunk.iterrows():
            # Extract the fields safely
            cves = safe_get(row, cols_lower, resolved.get("cves") or "")
            severity = safe_get(row, cols_lower, resolved.get("severity") or "")
            jfrog_sev = safe_get(row, cols_lower, resolved.get("jfrog_severity") or "")
            cvss = safe_get(row, cols_lower, resolved.get("cvss_v3") or "")
            cwe = safe_get(row, cols_lower, resolved.get("cwe") or "")
            fix_ver = safe_get(row, cols_lower, resolved.get("fix_version") or "")

            # Normalize severity to count categories
            sev_lower = severity.strip().lower()
            if "critical" in sev_lower:
                local_counts["critical"] += 1
            elif "high" in sev_lower:
                local_counts["high"] += 1
            elif "medium" in sev_lower:
                local_counts["medium"] += 1
            elif "low" in sev_lower:
                local_counts["low"] += 1
            else:
                # If severity is blank / unknown, attempt to infer from jfrog severity
                jf_lower = jfrog_sev.strip().lower()
                if "critical" in jf_lower:
                    local_counts["critical"] += 1
                elif "high" in jf_lower:
                    local_counts["high"] += 1
                elif "medium" in jf_lower:
                    local_counts["medium"] += 1
                elif "low" in jf_lower:
                    local_counts["low"] += 1
                # else ignore unknown severities for counting

            chunk_details.append((cves, severity, jfrog_sev, cvss, cwe, fix_ver))

        return chunk_details, local_counts

    # read file size and choose strategy
    try:
        file_size = os.path.getsize(path)
    except OSError:
        file_size = 0

    # If large file (> 10 MB), use chunked reading and threadpool
    use_chunking = file_size > (10 * 1024 * 1024) or True  # default True to be safe for many files

    if use_chunking:
        # read in chunks and process chunks using ThreadPoolExecutor
        try:
            reader = pd.read_csv(path, dtype=str, chunksize=CHUNK_SIZE, sep=None, engine="python", encoding="utf-8")
            futures = []
            with ThreadPoolExecutor(max_workers=THREADS_PER_WORKER) as tpool:
                for chunk in reader:
                    futures.append(tpool.submit(process_chunk, chunk))

                for fut in as_completed(futures):
                    try:
                        c_details, c_counts = fut.result()
                        details_acc.extend(c_details)
                        for k in severity_counts:
                            severity_counts[k] += c_counts.get(k, 0)
                    except Exception as ex:
                        logger.warning(f"Chunk processing failed for {filename}: {ex}")

        except Exception as e:
            # Fallback: try reading entire CSV at once
            logger.warning(f"Chunked read failed for {filename}: {e} — falling back to full read.")
            try:
                df_all = pd.read_csv(path, dtype=str, sep=None, engine="python", encoding="utf-8")
                c_details, c_counts = process_chunk(df_all)
                details_acc.extend(c_details)
                for k in severity_counts:
                    severity_counts[k] += c_counts.get(k, 0)
            except Exception as e2:
                logger.error(f"Failed to parse file {filename} on fallback: {e2}")
                return {"filename": filename, "counts": severity_counts, "details": []}
    else:
        # direct full read
        try:
            df_all = pd.read_csv(path, dtype=str, sep=None, engine="python", encoding="utf-8")
            c_details, c_counts = process_chunk(df_all)
            details_acc.extend(c_details)
            for k in severity_counts:
                severity_counts[k] += c_counts.get(k, 0)
        except Exception as e:
            logger.error(f"Failed to read file {filename}: {e}")
            return {"filename": filename, "counts": severity_counts, "details": []}

    logger.info(f"Worker done: {filename} counts={severity_counts} rows={len(details_acc)}")
    return {"filename": filename, "counts": severity_counts, "details": details_acc}


# ------------
# MAIN: coordinate multiprocessing and write XLSX
# ------------
def consolidate_all_csvs(base_dir: str, out_xlsx: str):
    base_dir = os.path.abspath(base_dir)
    logger.info(f"Scanning for CSVs in: {base_dir}")
    patterns = [os.path.join(base_dir, "*.csv"), os.path.join(base_dir, "*.CSV")]
    files = []
    for p in patterns:
        files.extend(glob.glob(p))

    files = sorted(list(set(files)))
    if not files:
        logger.error("No CSV files found.")
        return

    logger.info(f"Found {len(files)} CSV files. Processing with {NUM_PROCESSES} processes...")

    results = []
    # Use ProcessPoolExecutor for multiprocessing
    with ProcessPoolExecutor(max_workers=NUM_PROCESSES) as pool:
        future_to_file = {pool.submit(process_csv_file, f): f for f in files}
        for fut in as_completed(future_to_file):
            fpath = future_to_file[fut]
            try:
                res = fut.result()
                results.append(res)
            except Exception as exc:
                logger.error(f"Processing failed for {fpath}: {exc}")

    # Now write consolidated XLSX
    logger.info(f"Writing consolidated XLSX to: {out_xlsx}")
    # We'll create a single sheet and append blocks; use openpyxl engine.
    with pd.ExcelWriter(out_xlsx, engine="openpyxl", mode="w") as writer:
        sheet_name = "Consolidated"
        # Create empty DataFrame to start and write as placeholder to ensure sheet exists
        pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    # We will use openpyxl to append rows (pandas ExcelWriter doesn't append easily)
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(out_xlsx)
    ws = wb.active
    ws.title = "Consolidated"

    start_row = 1

    for res in results:
        fname = res.get("filename", "UNKNOWN")
        counts = res.get("counts", {})
        details = res.get("details", [])

        # Write filename as a bold header row (we'll just set value; formatting can be added)
        ws.cell(row=start_row, column=1, value=f"File: {fname}")
        start_row += 1

        # Write counts row
        c_crit = counts.get("critical", 0)
        c_high = counts.get("high", 0)
        c_med = counts.get("medium", 0)
        c_low = counts.get("low", 0)
        counts_text = f"Critical: {c_crit}    High: {c_high}    Medium: {c_med}    Low: {c_low}"
        ws.cell(row=start_row, column=1, value=counts_text)
        start_row += 1

        # Header for detail table
        headers = ["CVEs", "Severity", "Jfrog Severity", "CVSS v3", "Cwe", "Fix Version"]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=start_row, column=col_idx, value=h)
        start_row += 1

        # Write detail rows
        for detail in details:
            # detail is tuple of 6 fields
            for col_idx, val in enumerate(detail, start=1):
                ws.cell(row=start_row, column=col_idx, value=val)
            start_row += 1

        # blank separation row
        start_row += 1

    # Optionally auto-fit column widths (basic)
    for col in range(1, 7):
        max_len = 0
        for cell in ws[get_column_letter(col)]:
            if cell.value:
                l = len(str(cell.value))
                if l > max_len:
                    max_len = l
        # Add a little padding
        adjusted = min(100, max_len + 2)
        ws.column_dimensions[get_column_letter(col)].width = adjusted

    wb.save(out_xlsx)
    logger.info("Consolidation complete.")


# ------------
# CLI
# ------------
if __name__ == "__main__":
    # Accept env override or CLI? For now, simple constants; change above or set environment variable
    base_dir = os.environ.get("COMP_BASE_OUTPUT", BASE_OUTPUT_DIR)
    out_xlsx = os.environ.get("COMP_CONSOLIDATED_XLSX", OUTPUT_XLSX)

    consolidate_all_csvs(base_dir, out_xlsx)
