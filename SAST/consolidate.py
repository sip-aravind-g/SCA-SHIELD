#!/usr/bin/env python3
"""
Hybrid Multiprocessing + Threading Image Scanner Individual Scan Reports Consolidated Script
Author: Aravind G | Hybrid Cloud DevSecOps Senior Cloud Engineer
Email: ezsecops@hpe.com

Purpose:
    - Scan path: ../components/1.5.4/OLD/output/
    - Each IMAGES scan file contains CVE data
    - Consolidates into One Unified CVE Report
"""
import os
import csv
import logging
import threading
import multiprocessing
from queue import Queue
from collections import Counter
from openpyxl import Workbook

#CONFIG--------------
INPUT_DIR = "../components/1.5.4/OLD/output/"
OUTPUT_FILE = "../components/1.5.4/OLD/output/o1.xlsx"
MAX_PROCESSES = min(4, multiprocessing.cpu_count())
THREADS_PER_PROCESS = 5

REQUIRED_COLUMNS = [
    "cves",
    "severity",
    "jfrog severity",
    "cvss v3",
    "cwe",
    "fix version",
]

#LOGGING---------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(processName)s] %(levelname)s - %(message)s",
)

#UTILITIES---------------
def normalize_headers(headers):
    return {h.strip().lower(): idx for idx, h in enumerate(headers)}


def split_cves(raw):
    if not raw:
        return []
    raw = raw.replace(";", ",")
    return [
        cve.strip()
        for cve in raw.split(",")
        if cve.strip().startswith("CVE-")
    ]


def safe_get(row, col_map, col_name):
    idx = col_map.get(col_name)
    return row[idx].strip() if idx is not None and idx < len(row) else ""


#CSV PROCESSOR----------------
def process_csv_file(csv_path):
    filename = os.path.basename(csv_path)
    severity_counts = Counter()
    details = []
    seen_cves = set()

    try:
        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            if not headers:
                logging.warning(f"{filename}: Empty CSV")
                return filename, severity_counts, details

            col_map = normalize_headers(headers)

            for col in REQUIRED_COLUMNS:
                if col not in col_map:
                    logging.warning(f"{filename}: Missing column '{col}'")

            for row in reader:
                raw_cves = safe_get(row, col_map, "cves")
                cve_list = split_cves(raw_cves)

                severity = safe_get(row, col_map, "severity")
                jfrog_sev = safe_get(row, col_map, "jfrog severity")
                cvss = safe_get(row, col_map, "cvss v3")
                cwe = safe_get(row, col_map, "cwe")
                fix_ver = safe_get(row, col_map, "fix version")

                for cve in cve_list:
                    if cve in seen_cves:
                        continue

                    seen_cves.add(cve)

                    sev_lower = severity.lower()
                    if "critical" in sev_lower:
                        severity_counts["Critical"] += 1
                    elif "high" in sev_lower:
                        severity_counts["High"] += 1
                    elif "medium" in sev_lower:
                        severity_counts["Medium"] += 1
                    elif "low" in sev_lower:
                        severity_counts["Low"] += 1

                    details.append([
                        cve,
                        severity,
                        jfrog_sev,
                        cvss,
                        cwe,
                        fix_ver,
                    ])

    except Exception as e:
        logging.error(f"Failed processing {filename}: {e}")

    return filename, severity_counts, details


#THREAD WORKER---------------
def thread_worker(queue, results, lock):
    while True:
        try:
            csv_path = queue.get_nowait()
        except:
            return

        result = process_csv_file(csv_path)

        with lock:
            results.append(result)

        queue.task_done()


#PROCESS WORKER--------------
def process_worker(file_queue, results, lock):
    threads = []
    for _ in range(THREADS_PER_PROCESS):
        t = threading.Thread(
            target=thread_worker,
            args=(file_queue, results, lock),
            daemon=True,
        )
        t.start()
        threads.append(t)

    for t in threads:
        t.join()


#EXCEL WRITER-------------
def write_excel(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "Xray Consolidated Report"

    row = 1

    for filename, counts, details in sorted(results):
        ws.cell(row=row, column=1, value=filename)
        row += 1

        for sev in ["Critical", "High", "Medium", "Low"]:
            ws.cell(row=row, column=1, value=sev)
            ws.cell(row=row, column=2, value=counts.get(sev, 0))
            row += 1

        row += 1

        headers = ["CVEs", "Severity", "Jfrog Severity", "CVSS v3", "Cwe", "Fix Version"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)

        row += 1

        for d in details:
            for col, val in enumerate(d, 1):
                ws.cell(row=row, column=col, value=val)
            row += 1

        row += 2

    wb.save(OUTPUT_FILE)
    logging.info(f"Excel report generated: {OUTPUT_FILE}")


#MAIN---------------
def main():
    if not os.path.isdir(INPUT_DIR):
        logging.error(f"Input directory not found: {INPUT_DIR}")
        return

    csv_files = [
        os.path.join(INPUT_DIR, f)
        for f in os.listdir(INPUT_DIR)
        if f.lower().endswith(".csv")
    ]

    if not csv_files:
        logging.warning("No CSV files found")
        return

    manager = multiprocessing.Manager()
    results = manager.list()
    lock = manager.Lock()
    file_queue = manager.Queue()

    for f in csv_files:
        file_queue.put(f)

    processes = []
    for _ in range(MAX_PROCESSES):
        p = multiprocessing.Process(
            target=process_worker,
            args=(file_queue, results, lock),
        )
        p.start()
        processes.append(p)

    for p in processes:
        p.join()

    write_excel(list(results))


if __name__ == "__main__":
    main()
